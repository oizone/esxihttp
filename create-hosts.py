#!/usr/bin/python3
import xlrd
import os
import re
import openpyxl
wb=openpyxl.load_workbook(filename='esxi-hosts.xlsx')
ws=wb["Hosts"]
#hosts=ws.tables["Hosts"]


#excel=xlrd.open_workbook('esxi-hosts.xlsx')
#sheet=excel.sheet_by_index(0)
bootcfg=open("cd/boot.cfg","r").read()
i = 2

#while i < sheet.nrows:
for i in ws.iter_rows(min_row=3):
    #ks='ks=nfs://{}/{}'.format(sheet.cell(0,1).value,i[0].value)
    #ks='ks=http://{}/{}/{}'.format(sheet.cell(0,1).value,i[7].value,i[0].value)
    #ks='ks=http://{}/{}'.format(sheet.cell(0,1).value,i[7].value)
    if not os.path.exists(i[0].value):
        os.mkdir(i[0].value)
    #if not os.path.exists(i[7].value):
    #    os.mkdir(i[7].value)
    #if not os.path.exists("config"):
    #    os.mkdir("config")
    #output=open("{}/{}".format(i[7].value,i[0].value),"w+")
    output=open("{}/ks.cfg".format(i[0].value),"w+")

    output.write('vmaccepteula\n')
 
    output.write('rootpw {}\n'.format(i[12].value))
    output.write('clearpart --alldrives --overwritevmfs\n')

    output.write('install {} --overwritevmfs --novmfsondisk\n'.format(i[8].value))
    output.write('keyboard Finnish\n')

    output.write("network --bootproto=static --device={} --ip={} --netmask={} --gateway={} --nameserver={} --hostname={} --vlanid={} --addvmportgroup=1\n".format(i[1].value,i[2].value,i[3].value,i[4].value,i[5].value,i[0].value,int(i[6].value)))

    output.write('reboot --noeject\n')

    output.write('%firstboot --interpreter=busybox\n')
    capacitydisks=i[9].value.split(',')
    for disk in capacitydisks:
        if disk!='':
            output.write('esxcli vsan storage tag add -t capacityFlash -d `esxcli storage core device list|grep -B 1 "Display Name:.*{}"|head -n 1`\n'.format(disk))

    output.write('esxcli network ip dns search add --domain={}\n'.format(i[11].value))

    output.write('esxcli network vswitch standard portgroup set -v {} -p "VM Network"\n'.format(int(i[6].value)))

    output.write('vim-cmd hostsvc/enable_ssh\n')

    ntps=i[10].value.split(',')
    for ntp in ntps:
        if ntp!='':
            output.write('echo "server {}" >> /etc/ntp.conf\n'.format(ntp))


    output.write('/sbin/chkconfig ntpd on\n')
    output.write('esxcli system settings advanced set -o /UserVars/HostClientCEIPOptIn -i {}\n'.format(i[13].value))
    if (i[14].value == "yes"):
        output.write('esxcli system settings advanced set -o "/Mem/AllocGuestLargePage" --int-value 0\n')
        output.write('esxcli system settings advanced set -o "/Mem/ShareForceSalting" --int-value 0\n')
    if (i[15].value == "yes"):
        output.write('esxcli vsan network ipv4 add -i {}\n'.format(i[1].value))
        output.write('esxcli vsan cluster new\n')
        output.write('esxcli vsan policy setdefault -c cluster -p "((\"hostFailuresToTolerate\" i0) (\"forceProvisioning\" i1))"\n')
        output.write('esxcli vsan policy setdefault -c vdisk -p "((\"hostFailuresToTolerate\" i0) (\"forceProvisioning\" i1))")"\n')
        output.write('esxcli vsan policy setdefault -c vmnamespace -p "((\"hostFailuresToTolerate\" i0) (\"forceProvisioning\" i1))"\n')
        output.write('esxcli vsan policy setdefault -c vmswap -p "((\"hostFailuresToTolerate\" i0) (\"forceProvisioning\" i1))"\n')
        output.write('esxcli vsan policy setdefault -c vmem -p "((\"hostFailuresToTolerate\" i0) (\"forceProvisioning\" i1))"\n')
    output.write('#esxcli network ip interface remove --interface-name=vmk1\n')
    output.write('#esxcli network vswitch standard portgroup remove --portgroup-name="iDRAC Network" --vswitch-name=vSwitchiDRACvusb\n')
    output.write('#esxcli network vswitch standard remove --vswitch-name=vSwitchiDRACvusb\n')
    output.write('reboot\n')
    output.close()
    #if os.path.exists("config/{}".format(i[0].value.split('.')[0][-8:])):
    #    os.remove("config/{}".format(i[0].value.split('.')[0][-8:]))
    #os.symlink("{}".format(i[7].value),"config/{}".format(i[0].value.split('.')[0][-8:]))
    
    boot=open("{}/boot.cfg".format(i[0].value),"w+")
    #boot=open("pxelinux.cfg/{}.boot.cfg".format(i[7].value),"w+")
    newboot=re.sub("/","",bootcfg,flags=re.M)
    newboot=re.sub(r'prefix=[^\n]*','prefix=cd/',newboot,flags=re.M)
    #newboot=re.sub(r'prefix=[^\n]*','prefix=http://{}/cd/'.format(sheet.cell(0,1).value),newboot,flags=re.M)
    newboot=re.sub(r'kernelopt=[^\n]*','kernelopt={}'.format(ks),newboot,flags=re.M)
    boot.write(newboot)
    boot.close()
    

    i += 1





 



 
 

